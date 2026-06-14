"""
============================================================
PHẦN MỀM QUẢN LÝ DỮ LIỆU SÁT HẠCH GPLX
============================================================
File      : ui/tab_ie.py
Mô tả     : Tab IE v1.0
            - Import: Chọn XML → tự đọc + kiểm tra
            - Nơi cư trú load từ DVHC (TENDAYDU)
            - Click label "Danh sách" → load dữ liệu SH từ DB
            - Export: 1 dòng gồm Kỳ SH + Hạng + KQ + Excel + XML
            - Bỏ CSV, PDF, ô tick đầy đủ cột, nhập tiêu đề
            - Fix lỗi xlsxwriter → dùng openpyxl
Tác giả   : [ThienTon]
Phiên bản : 1.1.0
Ngày tạo  : 2026-06-12
============================================================

Bố cục:
┌─────────────────────────────────────────────────────────────────┐
│  ╔══════════════ 📥 IMPORT ══════════════════════════════════╗  │
│  ║ [📂 Chọn XML] [path_____] [🗑️ Xóa DL SH] [💾 Lưu DB]   ║  │
│  ║ File: xxx | Kỳ SH: xxx | NLX: 125 | ✅ 0 lỗi             ║  │
│  ║ □ Bỏ qua lỗi   □ Cập nhật trùng                          ║  │
│  ║                                                            ║  │
│  ║ 📋 Danh sách dữ liệu Import  ← click để load từ DB       ║  │
│  ║ ┌──────────────────────────────────────────────────────┐  ║  │
│  ║ │STT│Mã HV│Họ Tên│NSinh│Nơi CT│Hạng│SBD│NộiDung│LT│MP│H│Đ│KQ│ ║
│  ║ │ 1 │DK01 │Ng A  │1990 │TP.HN │B2  │01 │LT+H+D │✓│✗│✓│✓│Đạt│ ║
│  ║ └──────────────────────────────────────────────────────┘  ║  │
│  ║ ⚠️ Kết quả kiểm tra (3 mục)                               ║  │
│  ║ ┌──────────────────────────────────────────────────────┐  ║  │
│  ║ │Dòng│Cột│Loại│Mức độ│Mô tả                           │  ║  │
│  ║ └──────────────────────────────────────────────────────┘  ║  │
│  ╚═══════════════════════════════════════════════════════════╝  │
│                          ↕ resize                                │
│  ╔══════════════ 📤 EXPORT ══════════════════════════════════╗  │
│  ║ Kỳ SH:[▼] Hạng:[▼] KQ:[▼] [📊 Xuất Excel] [📄 Xuất XML] ║  │
│  ║ 📝 Nhật ký: [log...]                                      ║  │
│  ╚═══════════════════════════════════════════════════════════╝  │
│  [████████████████████ progress ████████████████████████████████]│
└─────────────────────────────────────────────────────────────────┘
"""

import os
import xml.etree.ElementTree as ET
from pathlib  import Path
from typing   import Optional
from datetime import datetime

from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGroupBox,
    QSplitter,
    QPushButton,
    QLabel,
    QLineEdit,
    QComboBox,
    QCheckBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QFileDialog,
    QMessageBox,
    QProgressBar,
    QTextEdit,
    QSpacerItem,
    QSizePolicy,
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui  import QColor, QFont, QCursor

from services.ie_service     import (
    IEService,
    ImportFileError,
    ImportXMLError,
    ExportError,
)
from services.logger_service import LoggerService
from database.db_manager     import DatabaseManager


# ══════════════════════════════════════════════════════════
#  STYLESHEET
# ══════════════════════════════════════════════════════════

TAB_IE_STYLE = """
QGroupBox {
    font-size: 10pt; font-weight: bold; color: #1a237e;
    border: 1px solid #c5cae9; border-radius: 6px;
    margin-top: 10px; padding-top: 15px;
}
QGroupBox::title {
    subcontrol-origin: margin; left: 15px; padding: 0 8px;
}
QPushButton {
    padding: 7px 15px; border-radius: 5px;
    font-size: 10pt; font-weight: bold;
    border: none; min-width: 85px;
}
QPushButton#btnChooseFile   { background-color: #1a237e; color: white; }
QPushButton#btnChooseFile:hover { background-color: #283593; }
QPushButton#btnDeleteData   { background-color: #F44336; color: white; }
QPushButton#btnDeleteData:hover { background-color: #C62828; }
QPushButton#btnSaveDB       { background-color: #4CAF50; color: white; }
QPushButton#btnSaveDB:hover { background-color: #388E3C; }
QPushButton#btnExportExcel  { background-color: #1B5E20; color: white; }
QPushButton#btnExportExcel:hover { background-color: #2E7D32; }
QPushButton#btnExportXML    { background-color: #0D47A1; color: white; }
QPushButton#btnExportXML:hover { background-color: #1565C0; }
QPushButton:disabled { background-color: #bdbdbd; color: #757575; }

QLineEdit, QComboBox {
    padding: 6px 10px; border: 1px solid #c5cae9;
    border-radius: 4px; font-size: 10pt;
}
QLineEdit:focus { border: 2px solid #1a237e; }

QTableWidget {
    border: 1px solid #c5cae9; border-radius: 4px;
    gridline-color: #e0e0e0; font-size: 9pt;
    alternate-background-color: #fafafa;
}
QTableWidget::item:selected {
    background-color: #bbdefb; color: #0d47a1;
}
QHeaderView::section {
    background-color: #1a237e; color: #ffffff;
    padding: 5px; border: 1px solid #283593;
    font-size: 9pt; font-weight: bold;
}

QProgressBar {
    border: 1px solid #c5cae9; border-radius: 4px;
    text-align: center; font-size: 9pt; height: 20px;
}
QProgressBar::chunk { background-color: #4CAF50; border-radius: 3px; }

QLabel#lblClickable {
    font-size: 10pt; font-weight: bold; color: #1a237e;
    padding: 4px 0; text-decoration: underline;
}
QLabel#lblClickable:hover {
    color: #0D47A1;
}
"""

# Cột preview
PREVIEW_COLUMNS = [
    # (key_nguoi, key_hoso, header, width, center)
    ("_index",       None,            "STT",         45,  True),
    ("MA_DK",        None,            "Mã HV",       90,  False),
    (None,           None,            "Họ Và Tên",  180,  False),
    ("NGAY_SINH",    None,            "Ngày Sinh",   85,  True),
    (None,           None,            "Nơi cư trú", 160,  False),  # load từ DVHC
    (None,           "HANG_GPLX",     "Hạng GPLX",   70,  True),
    (None,           "SO_BAO_DANH",   "SBD",         50,  True),
    (None,           "NOI_DUNG_SH",   "Nội dung SH",120,  False),  # load config_noidung
    (None,           "_LT",           "LT",          35,  True),
    (None,           "_MP",           "MP",          35,  True),
    (None,           "_H",            "H",           35,  True),
    (None,           "_D",            "Đ",           35,  True),
    (None,           "KET_QUA_SH",    "Kết Quả",     80,  True),
]

KQ_COLORS = {
    "Đạt": ("#1B5E20", "#E8F5E9"),
    "đạt": ("#1B5E20", "#E8F5E9"),
    "Dat": ("#1B5E20", "#E8F5E9"),
    "Không đạt": ("#B71C1C", "#FFEBEE"),
    "không đạt": ("#B71C1C", "#FFEBEE"),
    "Khong dat": ("#B71C1C", "#FFEBEE"),
}


# ══════════════════════════════════════════════════════════
#  CLASS TAB IE
# ══════════════════════════════════════════════════════════

class TabIE(QWidget):
    """Tab IE v2.2."""

    def __init__(
        self,
        ie_service : IEService,
        nguoi_lx_service,
        logger     : LoggerService,
        parent     : QWidget = None,
    ) -> None:
        super().__init__(parent)

        self.ie_svc = ie_service
        self.logger = logger
        self.db     = ie_service.db if hasattr(ie_service, 'db') else None

        # Cache
        self._noidung_map : dict = {}
        self._dvhc_map    : dict = {}  # {ma_dvhc: ten_day_du}
        self._load_noidung_config()
        self._load_dvhc_map()

        self.setStyleSheet(TAB_IE_STYLE)
        self._setup_ui()

    # ══════════════════════════════════════════════════════
    #  LOAD CACHE
    # ══════════════════════════════════════════════════════

    def _load_noidung_config(self) -> None:
        """Load bảng config_noidung."""
        self._noidung_map = {}
        if not self.db:
            return
        try:
            rows = self.db._fetchall(
                "SELECT * FROM config_noidung ORDER BY MA_NOIDUNG"
            )
            for r in rows:
                ma = str(r.get("MA_NOIDUNG", "")).strip()
                self._noidung_map[ma] = {
                    "LT": r.get("LT", 0),
                    "MP": r.get("MP", 0),
                    "H":  r.get("H", 0),
                    "D":  r.get("D", 0),
                    "MO_TA": r.get("MO_TA", ""),
                }
        except Exception:
            pass

    def _load_dvhc_map(self) -> None:
        """Load bảng dvhc → dict {ma_dvhc: TENDAYDU}."""
        self._dvhc_map = {}
        if not self.db:
            return
        try:
            rows = self.db._fetchall(
                "SELECT MA_DVHC, TENDAYDU FROM dvhc"
            )
            for r in rows:
                ma = str(r.get("MA_DVHC", "")).strip()
                self._dvhc_map[ma] = r.get("TENDAYDU", "") or ""
        except Exception:
            pass

    def _get_noidung_display(self, ma_nd: str) -> dict:
        """Lấy mô tả + LT/MP/H/D từ config_noidung."""
        ma = str(ma_nd or "").strip()
        if ma in self._noidung_map:
            nd = self._noidung_map[ma]
            return {
                "MO_TA": nd.get("MO_TA", ma),
                "LT": nd.get("LT", 0) == 1,
                "MP": nd.get("MP", 0) == 1,
                "H":  nd.get("H", 0) == 1,
                "D":  nd.get("D", 0) == 1,
            }
        return {"MO_TA": ma, "LT": False, "MP": False, "H": False, "D": False}

    def _get_noi_cu_tru(self, nguoi: dict) -> str:
        """
        Lấy Nơi cư trú từ DVHC.
        Tra cứu NOI_CT_MA_DVHC → TENDAYDU.
        Nếu không có thì dùng NOI_CT gốc.
        """
        # Thử tra từ mã DVHC
        ma_dvhc = str(nguoi.get("NOI_CT_MA_DVHC", "")).strip()
        if ma_dvhc and ma_dvhc in self._dvhc_map:
            return self._dvhc_map[ma_dvhc]

        # Fallback: dùng NOI_CT gốc
        return nguoi.get("NOI_CT", "") or ""

    # ══════════════════════════════════════════════════════
    #  SETUP UI
    # ══════════════════════════════════════════════════════

    def _setup_ui(self) -> None:
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(0)

        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self._create_import_section())
        splitter.addWidget(self._create_export_section())
        splitter.setSizes([620, 220])
        splitter.setCollapsible(0, False)
        splitter.setCollapsible(1, False)

        main_layout.addWidget(splitter, 1)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

    # ══════════════════════════════════════════════════════
    #  PHẦN IMPORT
    # ══════════════════════════════════════════════════════

    def _create_import_section(self) -> QGroupBox:
        group  = QGroupBox("📥  IMPORT - Nhập dữ liệu từ file XML")
        layout = QVBoxLayout(group)
        layout.setSpacing(6)
        layout.setContentsMargins(8, 18, 8, 8)

        # Dòng 1: Toolbar
        layout.addLayout(self._create_import_toolbar())

        # Dòng 2: Info + check status
        layout.addLayout(self._create_file_info_row())

        # Dòng 3: Tùy chọn
        layout.addLayout(self._create_import_options_row())

        # Splitter: Preview + Error
        import_split = QSplitter(Qt.Vertical)
        import_split.addWidget(self._create_preview_widget())
        import_split.addWidget(self._create_error_widget())
        import_split.setSizes([380, 100])
        layout.addWidget(import_split, 1)

        return group

    def _create_import_toolbar(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(6)

        self.btn_choose_file = QPushButton("📂 Chọn XML")
        self.btn_choose_file.setObjectName("btnChooseFile")
        self.btn_choose_file.setToolTip("Chọn file XML → Tự động đọc + kiểm tra")
        self.btn_choose_file.clicked.connect(self._on_choose_file)
        layout.addWidget(self.btn_choose_file)

        self.txt_file_path = QLineEdit()
        self.txt_file_path.setPlaceholderText("Chọn file XML → tự động đọc + kiểm tra...")
        self.txt_file_path.setReadOnly(True)
        layout.addWidget(self.txt_file_path, 1)

        layout.addSpacerItem(QSpacerItem(10, 0, QSizePolicy.Fixed))

        self.btn_delete_data = QPushButton("🗑️ Xóa DL sát hạch")
        self.btn_delete_data.setObjectName("btnDeleteData")
        self.btn_delete_data.setToolTip("Xóa: ky_sh + nguoi_lx + ho_so_sh + nhap_kqsh + log")
        self.btn_delete_data.clicked.connect(self._on_delete_all_data)
        layout.addWidget(self.btn_delete_data)

        self.btn_save_db = QPushButton("💾 Lưu DB")
        self.btn_save_db.setObjectName("btnSaveDB")
        self.btn_save_db.setEnabled(False)
        self.btn_save_db.clicked.connect(self._on_save_to_db)
        layout.addWidget(self.btn_save_db)

        return layout

    def _create_file_info_row(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(12)

        self.lbl_file_info = QLabel("File: ---")
        self.lbl_file_info.setStyleSheet("font-size: 9pt; color: #555;")
        layout.addWidget(self.lbl_file_info)

        self.lbl_ky_sh = QLabel("Kỳ SH: ---")
        self.lbl_ky_sh.setStyleSheet("font-size: 9pt; color: #555;")
        layout.addWidget(self.lbl_ky_sh)

        self.lbl_total = QLabel("NLX: ---")
        self.lbl_total.setStyleSheet("font-size: 9pt; color: #555;")
        layout.addWidget(self.lbl_total)

        self.lbl_check_status = QLabel("")
        self.lbl_check_status.setStyleSheet("font-size: 9pt; font-weight: bold;")
        layout.addWidget(self.lbl_check_status)

        layout.addStretch()
        return layout

    def _create_import_options_row(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(15)

        self.chk_skip_errors = QCheckBox("Bỏ qua dòng lỗi")
        self.chk_skip_errors.setChecked(True)
        self.chk_skip_errors.setStyleSheet("font-size: 9pt;")
        layout.addWidget(self.chk_skip_errors)

        self.chk_update_existing = QCheckBox("Cập nhật trùng MA_DK")
        self.chk_update_existing.setChecked(False)
        self.chk_update_existing.setStyleSheet("font-size: 9pt;")
        layout.addWidget(self.chk_update_existing)

        layout.addStretch()

        self.lbl_import_status = QLabel("")
        self.lbl_import_status.setStyleSheet("font-size: 9pt; font-weight: bold;")
        layout.addWidget(self.lbl_import_status)

        return layout

    # ──────────────────────────────────────────────────────
    #  PREVIEW: Label clickable + bảng
    # ──────────────────────────────────────────────────────

    def _create_preview_widget(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        # Label clickable
        self.lbl_preview_title = QLabel("📋 Danh sách dữ liệu Import")
        self.lbl_preview_title.setObjectName("lblClickable")
        self.lbl_preview_title.setCursor(QCursor(Qt.PointingHandCursor))
        self.lbl_preview_title.setToolTip(
            "Click để load dữ liệu sát hạch từ database"
        )
        self.lbl_preview_title.mousePressEvent = self._on_click_load_db
        layout.addWidget(self.lbl_preview_title)

        # Bảng
        headers = [c[2] for c in PREVIEW_COLUMNS]
        self.tbl_preview = QTableWidget()
        self.tbl_preview.setColumnCount(len(headers))
        self.tbl_preview.setHorizontalHeaderLabels(headers)
        self.tbl_preview.setAlternatingRowColors(True)
        self.tbl_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_preview.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_preview.verticalHeader().setDefaultSectionSize(28)
        self.tbl_preview.verticalHeader().setVisible(False)

        for i, (_, _, _, w, _) in enumerate(PREVIEW_COLUMNS):
            self.tbl_preview.setColumnWidth(i, w)
        self.tbl_preview.horizontalHeader().setStretchLastSection(True)

        layout.addWidget(self.tbl_preview)
        return widget

    def _create_error_widget(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        self.lbl_error_title = QLabel("⚠️ Kết quả kiểm tra")
        self.lbl_error_title.setStyleSheet(
            "font-size: 9pt; font-weight: bold; color: #E65100; padding: 2px 0;"
        )
        layout.addWidget(self.lbl_error_title)

        self.tbl_errors = QTableWidget()
        self.tbl_errors.setColumnCount(5)
        self.tbl_errors.setHorizontalHeaderLabels(["Dòng","Cột","Loại","Mức độ","Mô tả"])
        self.tbl_errors.setAlternatingRowColors(True)
        self.tbl_errors.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_errors.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_errors.verticalHeader().setDefaultSectionSize(26)
        h = self.tbl_errors.horizontalHeader()
        for i in range(4):
            h.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.Stretch)
        layout.addWidget(self.tbl_errors)

        return widget

    # ══════════════════════════════════════════════════════
    #  PHẦN EXPORT (1 dòng duy nhất)
    # ══════════════════════════════════════════════════════

    def _create_export_section(self) -> QGroupBox:
        group  = QGroupBox("📤  EXPORT - Xuất dữ liệu")
        layout = QVBoxLayout(group)
        layout.setSpacing(6)
        layout.setContentsMargins(8, 18, 8, 8)

        # Dòng duy nhất: Filter + Nút xuất
        layout.addLayout(self._create_export_row())

        # Log
        lbl_log = QLabel("📝 Nhật ký:")
        lbl_log.setStyleSheet("font-size: 9pt; font-weight: bold; color: #1a237e;")
        layout.addWidget(lbl_log)

        self.txt_export_log = QTextEdit()
        self.txt_export_log.setReadOnly(True)
        self.txt_export_log.setMaximumHeight(80)
        self.txt_export_log.setPlaceholderText("Nhật ký xuất dữ liệu...")
        self.txt_export_log.setStyleSheet(
            "font-family: Consolas, monospace; font-size: 9pt; "
            "background: #fafafa; border: 1px solid #c5cae9; border-radius: 4px;"
        )
        layout.addWidget(self.txt_export_log, 1)

        return group

    def _create_export_row(self) -> QHBoxLayout:
        """
        1 dòng duy nhất:
        [Kỳ SH ▼] [Hạng ▼] [KQ ▼] [📊 Xuất Excel] [📄 Xuất XML]
        """
        layout = QHBoxLayout()
        layout.setSpacing(8)

        # Kỳ SH
        layout.addWidget(QLabel("Kỳ SH:"))
        self.cmb_export_kysh = QComboBox()
        self.cmb_export_kysh.setMinimumWidth(130)
        self.cmb_export_kysh.addItem("-- Tất cả --", "")
        self._load_ky_sh_combo()
        layout.addWidget(self.cmb_export_kysh)

        # Hạng
        layout.addWidget(QLabel("Hạng:"))
        self.cmb_export_hang = QComboBox()
        self.cmb_export_hang.setMinimumWidth(80)
        self.cmb_export_hang.addItem("-- Tất cả --", "")
        for h in ["A1","A2","B1","B2","C","D","E","F"]:
            self.cmb_export_hang.addItem(h, h)
        layout.addWidget(self.cmb_export_hang)

        # KQ
        layout.addWidget(QLabel("KQ:"))
        self.cmb_export_kq = QComboBox()
        self.cmb_export_kq.setMinimumWidth(100)
        self.cmb_export_kq.addItem("-- Tất cả --", "")
        self.cmb_export_kq.addItem("Đạt", "Đạt")
        self.cmb_export_kq.addItem("Không đạt", "Không đạt")
        layout.addWidget(self.cmb_export_kq)

        layout.addSpacerItem(QSpacerItem(15, 0, QSizePolicy.Fixed))

        # Nút Xuất Excel
        self.btn_export_excel = QPushButton("📊 Xuất Excel")
        self.btn_export_excel.setObjectName("btnExportExcel")
        self.btn_export_excel.clicked.connect(self._on_export_excel)
        layout.addWidget(self.btn_export_excel)

        # Nút Xuất XML
        self.btn_export_xml = QPushButton("📄 Xuất XML")
        self.btn_export_xml.setObjectName("btnExportXML")
        self.btn_export_xml.clicked.connect(self._on_export_xml)
        layout.addWidget(self.btn_export_xml)

        layout.addStretch()

        # Label kết quả
        self.lbl_export_result = QLabel("")
        self.lbl_export_result.setStyleSheet("font-size: 10pt; font-weight: bold;")
        layout.addWidget(self.lbl_export_result)

        return layout

    # ══════════════════════════════════════════════════════
    #  IMPORT: CHỌN FILE → TỰ ĐỘNG ĐỌC + KIỂM TRA
    # ══════════════════════════════════════════════════════

    def _on_choose_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file dữ liệu", "",
            "XML Files (*.xml);;"
            "Excel Files (*.xlsx);;"
            "Tất cả (*.xml *.xlsx *.csv)",
        )
        if not path:
            return

        self.txt_file_path.setText(path)
        self._reset_import_ui()
        self.ie_svc.reset_import()

        try:
            self._show_progress(True)

            ext = Path(path).suffix.lower()

            # Bước 1: Đọc
            if ext == ".xml":
                self.ie_svc.doc_file_xml(path)
            else:
                self.ie_svc.doc_file(path)

            self._update_file_info_after_read(path)

            # Bước 2: Kiểm tra lỗi
            errors = self.ie_svc.kiem_tra_xml() if ext == ".xml" else []

            # Bước 3: Hiển thị TOÀN BỘ danh sách
            self._show_full_preview_import()

            # Bước 4: Hiển thị lỗi
            self._show_errors(errors)
            self._update_check_status(errors)

            self._show_progress(False)

            self.lbl_preview_title.setText(
                f"📋 Danh sách dữ liệu Import ({self.ie_svc.row_count} NLX)"
            )

            self._notify_parent(
                f"📖 Đọc + kiểm tra: {Path(path).name} "
                f"({self.ie_svc.row_count} NLX)"
            )

        except (ImportFileError, ImportXMLError) as exc:
            self._show_progress(False)
            QMessageBox.warning(self, "Lỗi đọc file", str(exc))
        except Exception as exc:
            self._show_progress(False)
            QMessageBox.critical(self, "Lỗi", f"Lỗi:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  CLICK LABEL → LOAD DỮ LIỆU TỪ DB
    # ══════════════════════════════════════════════════════

    def _on_click_load_db(self, event) -> None:
        """
        Click vào label "Danh sách dữ liệu Import"
        → Load dữ liệu sát hạch từ database (JOIN 3 bảng).
        """
        if not self.db:
            return

        try:
            self._show_progress(True)

            rows = self.db._fetchall("""
                SELECT n.SO_TT, n.MA_DK, n.HO_VA_TEN, n.NGAY_SINH,
                       n.NOI_CT, n.NOI_CT_MA_DVHC,
                       h.HANG_GPLX, h.SO_BAO_DANH, h.NOI_DUNG_SH,
                       h.KET_QUA_SH
                FROM nguoi_lx n
                LEFT JOIN ho_so_sh h ON n.MA_DK = h.MA_DK
                ORDER BY n.SO_TT
                LIMIT 10000
            """)

            if not rows:
                self._show_progress(False)
                self.lbl_preview_title.setText(
                    "📋 Danh sách dữ liệu sát hạch (0 bản ghi)"
                )
                self.tbl_preview.setRowCount(0)
                QMessageBox.information(
                    self, "Thông báo",
                    "Không có dữ liệu sát hạch trong database."
                )
                return

            # Hiển thị
            total = len(rows)
            self.tbl_preview.setRowCount(total)

            for ri, row in enumerate(rows):
                ma_nd = str(row.get("NOI_DUNG_SH", "")).strip()
                nd_info = self._get_noidung_display(ma_nd)

                # Nơi cư trú từ DVHC
                ma_dvhc = str(row.get("NOI_CT_MA_DVHC", "")).strip()
                noi_ct = (
                    self._dvhc_map.get(ma_dvhc, "")
                    or row.get("NOI_CT", "")
                    or ""
                )

                for ci, (key_n, key_h, header, w, center) in enumerate(PREVIEW_COLUMNS):
                    value = ""
                    fg = None
                    bg = None
                    bold = False

                    if header == "STT":
                        value = str(ri + 1)
                    elif header == "Mã HV":
                        value = row.get("MA_DK", "")
                    elif header == "Họ Và Tên":
                        value = row.get("HO_VA_TEN", "")
                    elif header == "Ngày Sinh":
                        value = row.get("NGAY_SINH", "")
                    elif header == "Nơi cư trú":
                        value = noi_ct
                    elif header == "Hạng GPLX":
                        value = row.get("HANG_GPLX", "")
                    elif header == "SBD":
                        value = row.get("SO_BAO_DANH", "")
                    elif header == "Nội dung SH":
                        value = nd_info["MO_TA"]
                    elif header in ("LT", "MP", "H", "Đ"):
                        key_map = {"LT":"LT","MP":"MP","H":"H","Đ":"D"}
                        val = nd_info.get(key_map.get(header, ""), False)
                        value = "✓" if val else "✗"
                        fg = "#1B5E20" if val else "#BDBDBD"
                        bg = "#E8F5E9" if val else None
                        bold = val
                    elif header == "Kết Quả":
                        value = row.get("KET_QUA_SH", "")
                        colors = KQ_COLORS.get(str(value).strip())
                        if colors:
                            fg, bg = colors
                            bold = True

                    item = QTableWidgetItem(str(value or ""))
                    if center:
                        item.setTextAlignment(Qt.AlignCenter)
                    if fg:
                        item.setForeground(QColor(fg))
                    if bg:
                        item.setBackground(QColor(bg))
                    if bold:
                        f = QFont(); f.setBold(True); item.setFont(f)
                    self.tbl_preview.setItem(ri, ci, item)

            self._show_progress(False)

            self.lbl_preview_title.setText(
                f"📋 Danh sách dữ liệu sát hạch ({total} bản ghi từ DB)"
            )
            self._notify_parent(f"📋 Load {total} bản ghi từ DB")

        except Exception as exc:
            self._show_progress(False)
            QMessageBox.warning(self, "Lỗi", f"Lỗi load DB:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  PREVIEW IMPORT (từ file XML)
    # ══════════════════════════════════════════════════════

    def _show_full_preview_import(self) -> None:
        """Hiển thị TOÀN BỘ danh sách từ file import."""
        nguoi_list = self.ie_svc.xml_nguoi_lx
        hoso_list  = self.ie_svc.xml_ho_so

        if not nguoi_list:
            return

        total = len(nguoi_list)
        self.tbl_preview.setRowCount(total)

        for ri in range(total):
            nguoi = nguoi_list[ri]
            hoso  = hoso_list[ri] if ri < len(hoso_list) else {}

            ma_nd = str(hoso.get("NOI_DUNG_SH", "")).strip()
            nd_info = self._get_noidung_display(ma_nd)

            # Nơi cư trú từ DVHC
            noi_ct = self._get_noi_cu_tru(nguoi)

            for ci, (key_n, key_h, header, w, center) in enumerate(PREVIEW_COLUMNS):
                value = ""
                fg = None
                bg = None
                bold = False

                if header == "STT":
                    value = str(ri + 1)
                elif header == "Mã HV":
                    value = nguoi.get("MA_DK", "")
                elif header == "Họ Và Tên":
                    value = nguoi.get("HO_VA_TEN", "")
                elif header == "Ngày Sinh":
                    value = nguoi.get("NGAY_SINH", "")
                elif header == "Nơi cư trú":
                    value = noi_ct
                elif header == "Hạng GPLX":
                    value = hoso.get("HANG_GPLX", "")
                elif header == "SBD":
                    value = hoso.get("SO_BAO_DANH", "")
                elif header == "Nội dung SH":
                    value = nd_info["MO_TA"]
                elif header in ("LT", "MP", "H", "Đ"):
                    key_map = {"LT":"LT","MP":"MP","H":"H","Đ":"D"}
                    val = nd_info.get(key_map.get(header, ""), False)
                    value = "✓" if val else "✗"
                    fg = "#1B5E20" if val else "#BDBDBD"
                    bg = "#E8F5E9" if val else None
                    bold = val
                elif header == "Kết Quả":
                    value = hoso.get("KET_QUA_SH", "")
                    colors = KQ_COLORS.get(str(value).strip())
                    if colors:
                        fg, bg = colors
                        bold = True
                elif key_n:
                    value = nguoi.get(key_n, "")
                elif key_h and not key_h.startswith("_"):
                    value = hoso.get(key_h, "")

                item = QTableWidgetItem(str(value or ""))
                if center:
                    item.setTextAlignment(Qt.AlignCenter)
                if fg:
                    item.setForeground(QColor(fg))
                if bg:
                    item.setBackground(QColor(bg))
                if bold:
                    f = QFont(); f.setBold(True); item.setFont(f)
                self.tbl_preview.setItem(ri, ci, item)

    # ══════════════════════════════════════════════════════
    #  ERROR, STATUS, RESET
    # ══════════════════════════════════════════════════════

    def _show_errors(self, errors: list[dict]) -> None:
        self.tbl_errors.setRowCount(len(errors))
        if not errors:
            self.lbl_error_title.setText("✅ Không có lỗi")
            self.lbl_error_title.setStyleSheet(
                "font-size: 9pt; font-weight: bold; color: #1B5E20;"
            )
            return

        self.lbl_error_title.setText(f"⚠️ Kết quả kiểm tra ({len(errors)} mục)")
        self.lbl_error_title.setStyleSheet(
            "font-size: 9pt; font-weight: bold; color: #E65100;"
        )

        for ri, err in enumerate(errors):
            it = QTableWidgetItem(str(err.get("row", "")))
            it.setTextAlignment(Qt.AlignCenter)
            self.tbl_errors.setItem(ri, 0, it)
            self.tbl_errors.setItem(ri, 1, QTableWidgetItem(err.get("column", "")))
            self.tbl_errors.setItem(ri, 2, QTableWidgetItem(err.get("type", "")))

            sev = err.get("severity", "error")
            it_s = QTableWidgetItem(sev.upper())
            it_s.setTextAlignment(Qt.AlignCenter)
            if sev == "error":
                it_s.setForeground(QColor("#B71C1C"))
                it_s.setBackground(QColor("#FFEBEE"))
                f = QFont(); f.setBold(True); it_s.setFont(f)
            else:
                it_s.setForeground(QColor("#E65100"))
                it_s.setBackground(QColor("#FFF3E0"))
            self.tbl_errors.setItem(ri, 3, it_s)
            self.tbl_errors.setItem(ri, 4, QTableWidgetItem(err.get("message", "")))

    def _update_check_status(self, errors: list[dict]) -> None:
        n_err  = sum(1 for e in errors if e.get("severity") == "error")
        n_warn = sum(1 for e in errors if e.get("severity") == "warning")

        if n_err == 0 and n_warn == 0:
            self.lbl_check_status.setText("✅ Không lỗi")
            self.lbl_check_status.setStyleSheet("font-size: 9pt; font-weight: bold; color: #1B5E20;")
            self.btn_save_db.setEnabled(True)
        elif n_err == 0:
            self.lbl_check_status.setText(f"⚠️ {n_warn} cảnh báo")
            self.lbl_check_status.setStyleSheet("font-size: 9pt; font-weight: bold; color: #E65100;")
            self.btn_save_db.setEnabled(True)
        else:
            self.lbl_check_status.setText(f"❌ {n_err} lỗi, {n_warn} cảnh báo")
            self.lbl_check_status.setStyleSheet("font-size: 9pt; font-weight: bold; color: #B71C1C;")
            self.btn_save_db.setEnabled(self.chk_skip_errors.isChecked())

    def _update_file_info_after_read(self, file_path: str) -> None:
        p = Path(file_path)
        size = p.stat().st_size
        s = f"{size/(1024*1024):.1f}MB" if size > 1024*1024 else f"{size/1024:.1f}KB"
        self.lbl_file_info.setText(f"File: {p.name} ({s})")
        self.lbl_total.setText(f"NLX: {self.ie_svc.file_info.get('total_rows', 0)}")

        ky = self.ie_svc.xml_ky_sh
        if ky and ky.get("MAKYSH"):
            t = f"Kỳ SH: {ky['MAKYSH']}"
            if ky.get("NGAYSH"): t += f" ({ky['NGAYSH']})"
            self.lbl_ky_sh.setText(t)
        else:
            self.lbl_ky_sh.setText("Kỳ SH: ---")

    def _reset_import_ui(self) -> None:
        self.btn_save_db.setEnabled(False)
        self.tbl_preview.setRowCount(0)
        self.tbl_errors.setRowCount(0)
        self.lbl_check_status.setText("")
        self.lbl_import_status.setText("")
        self.lbl_file_info.setText("File: ---")
        self.lbl_ky_sh.setText("Kỳ SH: ---")
        self.lbl_total.setText("NLX: ---")
        self.lbl_preview_title.setText("📋 Danh sách dữ liệu Import")

    # ══════════════════════════════════════════════════════
    #  XÓA TOÀN BỘ DỮ LIỆU SÁT HẠCH
    # ══════════════════════════════════════════════════════

    def _on_delete_all_data(self) -> None:
        if not self.db:
            return

        try:
            counts = {}
            for t in ["ky_sh","nguoi_lx","ho_so_sh","nhap_kqsh"]:
                r = self.db._fetchone(f"SELECT COUNT(*) as cnt FROM {t}")
                counts[t] = r["cnt"] if r else 0
            total = sum(counts.values())
        except Exception:
            total = 0; counts = {}

        if total == 0:
            QMessageBox.information(self, "Thông báo", "Không có dữ liệu để xóa.")
            return

        detail = "\n".join([f"  • {t}: {c}" for t,c in counts.items()])

        reply = QMessageBox.warning(
            self, "⚠️ XÓA TOÀN BỘ",
            f"XÓA TOÀN BỘ dữ liệu sát hạch?\n\n{detail}\n  • File log\n\n"
            f"TỔNG: {total} bản ghi\n\n⚠️ KHÔNG THỂ HOÀN TÁC!",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        try:
            self._show_progress(True)
            self.db._execute("DELETE FROM nhap_kqsh")
            self.db._execute("DELETE FROM ho_so_sh")
            self.db._execute("DELETE FROM nguoi_lx")
            self.db._execute("DELETE FROM ky_sh")
            self.db.commit()

            log_path = Path(self.logger.log_path)
            if log_path.exists():
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("")

            self._show_progress(False)
            self.logger.info(module="IE", action="XoaDL", detail=f"Xoa {total} ban ghi")

            QMessageBox.information(self, "OK", f"✅ Đã xóa {total} bản ghi!")
            self._reset_import_ui()
            self.ie_svc.reset_import()
            self._reload_ky_sh_combo()
            self._notify_parent("🗑️ Đã xóa toàn bộ DL sát hạch")

        except Exception as exc:
            self._show_progress(False)
            self.db.rollback()
            QMessageBox.critical(self, "Lỗi", f"Lỗi:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  LƯU DATABASE
    # ══════════════════════════════════════════════════════

    def _on_save_to_db(self) -> None:
        total = self.ie_svc.row_count
        if total == 0:
            QMessageBox.information(self, "Thông báo", "Không có dữ liệu.")
            return

        skip   = self.chk_skip_errors.isChecked()
        update = self.chk_update_existing.isChecked()

        reply = QMessageBox.question(
            self, "Xác nhận",
            f"Import {total} hồ sơ vào DB?\n\n"
            f"  • {'Bỏ qua lỗi' if skip else 'Dừng khi lỗi'}\n"
            f"  • {'Cập nhật trùng' if update else 'Bỏ qua trùng'}",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, total)
            self.progress_bar.setValue(0)

            result = self.ie_svc.luu_xml_database(skip_errors=skip, update_existing=update)
            self.progress_bar.setValue(total)

            msg = (
                f"📥 Import hoàn tất!\n\n"
                f"  Tổng: {result.get('total',0)}\n"
                f"  OK: {result.get('success',0)}\n"
                f"  Cập nhật: {result.get('updated',0)}\n"
                f"  Lỗi: {result.get('failed',0)}\n"
                f"  Bỏ qua: {result.get('skipped',0)}\n"
            )
            if result.get("ky_sh_saved"):
                msg += "\n  ✅ Kỳ SH đã lưu.\n"

            errs = result.get("errors", [])
            if errs:
                msg += "\n── Lỗi ──\n"
                for e in errs[:10]:
                    msg += f"  • {e}\n"

            QMessageBox.information(self, "Kết quả", msg)
            self._show_progress(False)

            self.lbl_import_status.setText(f"✅ {result.get('success',0)} OK")
            self.lbl_import_status.setStyleSheet("font-size: 9pt; font-weight: bold; color: #1B5E20;")

            self._reload_ky_sh_combo()
            self._notify_parent(f"✅ Import: {result.get('success',0)} OK")

        except Exception as exc:
            self._show_progress(False)
            QMessageBox.critical(self, "Lỗi", f"Lỗi:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  EXPORT EXCEL (dùng openpyxl thay xlsxwriter)
    # ══════════════════════════════════════════════════════

    def _on_export_excel(self) -> None:
        filters = self._get_export_filters()
        name = self.ie_svc.get_suggested_filename("xlsx", filters.get("hang_gplx",""))
        path, _ = QFileDialog.getSaveFileName(self, "Xuất Excel", name, "Excel (*.xlsx)")
        if not path:
            return

        try:
            self._show_progress(True)

            # Lấy dữ liệu
            data = self.ie_svc._get_export_data(**filters)

            if not data:
                self._show_progress(False)
                QMessageBox.information(self, "Thông báo", "Không có dữ liệu.")
                return

            # ── Dùng openpyxl thay vì xlsxwriter ──────────
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "DuLieu"

            # Style
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            cell_font = Font(name="Segoe UI", size=10)
            center_align = Alignment(horizontal="center", vertical="center")

            dat_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            kd_fill  = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

            # Cột xuất
            export_cols = [
                ("SO_TT",          "STT",          8),
                ("MA_DK",          "Mã ĐK",       15),
                ("HO_VA_TEN",      "Họ và tên",   30),
                ("NGAY_SINH",      "Ngày sinh",   14),
                ("SO_CMT",         "Số CCCD",     16),
                ("NOI_CT",         "Nơi cư trú", 35),
                ("HANG_GPLX",      "Hạng GPLX",  10),
                ("SO_BAO_DANH",    "SBD",         8),
                ("NOI_DUNG_SH",    "Nội dung SH",15),
                ("KQ_SH_LYTHUYET", "KQ LT",      10),
                ("KQ_SH_MOPHONG",  "KQ MP",      10),
                ("KQ_SH_HINH",     "KQ Hình",    10),
                ("KQ_SH_DUONG",    "KQ Đường",   10),
                ("KET_QUA_SH",     "Kết quả",    12),
            ]

            # Header
            for ci, (_, header, width) in enumerate(export_cols, 1):
                cell = ws.cell(row=1, column=ci, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                ws.column_dimensions[cell.column_letter].width = width

            # Data
            center_fields = {"SO_TT","HANG_GPLX","SO_BAO_DANH",
                             "KQ_SH_LYTHUYET","KQ_SH_MOPHONG",
                             "KQ_SH_HINH","KQ_SH_DUONG","KET_QUA_SH"}

            for ri, row_data in enumerate(data, 2):
                for ci, (field, _, _) in enumerate(export_cols, 1):
                    if field == "SO_TT":
                        value = ri - 1
                    elif field == "NOI_CT":
                        # Nơi cư trú từ DVHC
                        ma_dvhc = str(row_data.get("NOI_CT_MA_DVHC", "")).strip()
                        value = self._dvhc_map.get(ma_dvhc, row_data.get("NOI_CT", ""))
                    else:
                        value = row_data.get(field, "") or ""

                    cell = ws.cell(row=ri, column=ci, value=str(value))
                    cell.font = cell_font
                    cell.border = thin_border

                    if field in center_fields:
                        cell.alignment = center_align

                    # Tô màu kết quả
                    if field == "KET_QUA_SH":
                        v = str(value).strip().lower()
                        if v in ("đạt", "dat"):
                            cell.fill = dat_fill
                        elif "không" in v or "khong" in v:
                            cell.fill = kd_fill

            # Freeze panes
            ws.freeze_panes = "A2"

            # Auto filter
            ws.auto_filter.ref = f"A1:{ws.cell(1, len(export_cols)).column_letter}{len(data)+1}"

            # Lưu
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)

            self._show_progress(False)

            ts = datetime.now().strftime("%H:%M:%S")
            self.txt_export_log.append(
                f"[{ts}] ✅ EXCEL: {len(data)} hồ sơ → {Path(path).name}"
            )
            self.lbl_export_result.setText(f"✅ Xuất {len(data)} hồ sơ")
            self.lbl_export_result.setStyleSheet("font-size: 10pt; font-weight: bold; color: #1B5E20;")

            self.logger.info(module="IE", action="ExportExcel",
                             detail=f"{len(data)} ban ghi → {Path(path).name}")

            self._notify_parent(f"✅ Excel: {len(data)} hồ sơ")

            if QMessageBox.question(self, "OK", f"Xuất {len(data)} hồ sơ.\n\nMở file?") == QMessageBox.Yes:
                os.startfile(path)

        except ImportError:
            self._show_progress(False)
            QMessageBox.critical(
                self, "Thiếu thư viện",
                "Cần cài thư viện openpyxl:\n\n"
                "pip install openpyxl\n\n"
                "Chạy lệnh trên trong Terminal/CMD rồi thử lại."
            )
        except Exception as exc:
            self._show_progress(False)
            QMessageBox.critical(self, "Lỗi", f"Lỗi xuất Excel:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  EXPORT XML
    # ══════════════════════════════════════════════════════

    def _on_export_xml(self) -> None:
        """Xuất dữ liệu ra file XML theo cấu trúc SAT_HACH."""
        filters = self._get_export_filters()
        name = self.ie_svc.get_suggested_filename("xml", filters.get("hang_gplx",""))
        path, _ = QFileDialog.getSaveFileName(self, "Xuất XML", name, "XML (*.xml)")
        if not path:
            return

        try:
            self._show_progress(True)

            # Lấy dữ liệu
            data = self.ie_svc._get_export_data(**filters)
            if not data:
                self._show_progress(False)
                QMessageBox.information(self, "Thông báo", "Không có dữ liệu.")
                return

            # Lấy thông tin kỳ SH
            ma_ky = filters.get("ma_ky_sh", "")
            ky_sh_info = {}
            if ma_ky and self.db:
                row = self.db._fetchone(
                    "SELECT * FROM ky_sh WHERE MAKYSH = ?", (ma_ky,)
                )
                if row:
                    ky_sh_info = dict(row)

            # Build XML
            root = ET.Element("SAT_HACH")

            # HEADER
            header = ET.SubElement(root, "HEADER")
            ET.SubElement(header, "NGAY_XUAT").text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ET.SubElement(header, "SO_BAN_GHI").text = str(len(data))

            # DATA
            data_elem = ET.SubElement(root, "DATA")

            # KY_SH
            if ky_sh_info:
                ky_elem = ET.SubElement(data_elem, "KY_SH")
                for key, val in ky_sh_info.items():
                    ET.SubElement(ky_elem, key).text = str(val or "")

            # NGUOI_LXS
            nguoi_lxs = ET.SubElement(data_elem, "NGUOI_LXS")

            for row_data in data:
                ma_dk = row_data.get("MA_DK", "")
                if not ma_dk:
                    continue

                # Lấy đầy đủ nguoi_lx
                nguoi_full = self.db._fetchone(
                    "SELECT * FROM nguoi_lx WHERE MA_DK = ?", (ma_dk,)
                ) if self.db else {}

                hoso_full = self.db._fetchone(
                    "SELECT * FROM ho_so_sh WHERE MA_DK = ?", (ma_dk,)
                ) if self.db else {}

                nguoi_full = dict(nguoi_full) if nguoi_full else {}
                hoso_full  = dict(hoso_full)  if hoso_full  else {}

                # NGUOI_LX
                nlx = ET.SubElement(nguoi_lxs, "NGUOI_LX")
                nguoi_fields = [
                    "SO_TT","MA_DK","HO_TEN_DEM","TEN","HO_VA_TEN",
                    "GIOI_TINH","NGAY_SINH","MA_QUOC_TICH",
                    "NOI_CT","NOI_CT_MA_DVHC","NOI_CT_MA_DVQL","SO_CMT",
                ]
                for f in nguoi_fields:
                    ET.SubElement(nlx, f).text = str(nguoi_full.get(f, "") or "")

                # HO_SO
                hoso_elem = ET.SubElement(nlx, "HO_SO")
                hoso_fields = [
                    "MA_DK","SO_HO_SO","MA_KY_SH","SO_BAO_DANH",
                    "MA_CSDT","MA_TTSH","MA_SO_GTVT","GIAY_CNSK",
                    "HANG_GPLX","SO_GPLX_DA_CO","HANG_GPLX_DA_CO",
                    "DVQL_GPLX_DACO","NGAY_HH_GPLX_DACO",
                    "SO_NAM_LAIXE","SO_KM_ANTOAN","SO_GIAY_CNTN",
                    "SO_CCN","NOI_DUNG_SH","LY_DO_SH",
                    "KET_QUA_SH","KQ_SH_LYTHUYET","KQ_SH_MOPHONG",
                    "KQ_SH_HINH","KQ_SH_DUONG","GHI_CHU_SH",
                    "ANH_CHAN_DUNG","NGAY_TT_GPLX_DACO","MA_KHOA_HOC",
                    "SO_QD_SH","NGAY_QD_SH","NGUOI_QD_SH","CHAT_LUONG_ANH",
                ]
                for f in hoso_fields:
                    ET.SubElement(hoso_elem, f).text = str(hoso_full.get(f, "") or "")

            # Ghi file
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ")
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            tree.write(path, encoding="utf-8", xml_declaration=True)

            self._show_progress(False)

            ts = datetime.now().strftime("%H:%M:%S")
            self.txt_export_log.append(
                f"[{ts}] ✅ XML: {len(data)} hồ sơ → {Path(path).name}"
            )
            self.lbl_export_result.setText(f"✅ Xuất XML {len(data)} hồ sơ")
            self.lbl_export_result.setStyleSheet("font-size: 10pt; font-weight: bold; color: #1B5E20;")

            self.logger.info(module="IE", action="ExportXML",
                             detail=f"{len(data)} ban ghi → {Path(path).name}")

            self._notify_parent(f"✅ XML: {len(data)} hồ sơ")

            if QMessageBox.question(self, "OK", f"Xuất {len(data)} hồ sơ XML.\n\nMở file?") == QMessageBox.Yes:
                os.startfile(path)

        except Exception as exc:
            self._show_progress(False)
            QMessageBox.critical(self, "Lỗi", f"Lỗi xuất XML:\n{exc}")

    # ══════════════════════════════════════════════════════
    #  TIỆN ÍCH
    # ══════════════════════════════════════════════════════

    def _get_export_filters(self) -> dict:
        return {
            "ma_ky_sh":  self.cmb_export_kysh.currentData() or "",
            "hang_gplx": self.cmb_export_hang.currentData() or "",
            "ket_qua":   self.cmb_export_kq.currentData()   or "",
        }

    def _show_progress(self, show: bool) -> None:
        self.progress_bar.setVisible(show)
        if show:
            self.progress_bar.setRange(0, 0)

    def _load_ky_sh_combo(self) -> None:
        if not self.db: return
        try:
            rows = self.db._fetchall("SELECT MAKYSH, NGAYSH FROM ky_sh ORDER BY MAKYSH")
            for r in rows:
                lbl = r.get("MAKYSH", "")
                if r.get("NGAYSH"): lbl += f" ({r['NGAYSH']})"
                self.cmb_export_kysh.addItem(lbl, r["MAKYSH"])
        except Exception:
            pass

    def _reload_ky_sh_combo(self) -> None:
        self.cmb_export_kysh.clear()
        self.cmb_export_kysh.addItem("-- Tất cả --", "")
        self._load_ky_sh_combo()

    def _notify_parent(self, msg: str) -> None:
        p = self.parent()
        if p and hasattr(p, "show_status_message"):
            p.show_status_message(msg)

    # PUBLIC API
    def trigger_import_excel(self) -> None:
        self._on_choose_file()

    def trigger_import_csv(self) -> None:
        self._on_choose_file()

    def trigger_export_excel(self) -> None:
        self._on_export_excel()

    def trigger_export_csv(self) -> None:
        self._on_export_excel()

    def trigger_export_pdf(self) -> None:
        self._on_export_excel()